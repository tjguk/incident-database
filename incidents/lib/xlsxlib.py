import os, sys
import logging

logger = logging.getLogger(__package__)

import openpyxl

def mangle_sheet_name(sheet_name):
    "Sheet names cannot be longer than 31 chars"
    return sheet_name[:31]

def cells_in_range(range):
    for row in range:
        for cell in row:
            yield cell

def xlsx(data_iterator, spreadsheet_filepath, optimised=False, callback=None):
    """xlsx - put a dataset to an xlsx spreadsheet

    Parameters:
        an iterator which will supply [(Sheet Name, [Column Names / Types], [Rows]), ...]
        spreadsheet_filepath - full path to a spreadsheet
    """
    ROWSET_SIZE = 1000

    wb = openpyxl.Workbook(optimized_write=optimised)
    for sheet in list(wb.worksheets):
        wb.remove_sheet(sheet)

    for n_sheet, (sheet_name, headers, rowset) in enumerate(data_iterator):
        #
        # Create a new sheet to hold the rowset. Use the sheet name supplied,
        # adjusted as necessary to meet the constraints of Excel sheet names,
        # or create a default one if none is supplied.
        #
        if sheet_name:
            sheet_name = mangle_sheet_name(sheet_name)
        else:
            sheet_name = u"Sheet %d" % n_sheet
        ws = wb.create_sheet(title=sheet_name)
        ws.set_printer_settings("A4", ws.ORIENTATION_LANDSCAPE)
        yield "%s... " % sheet_name

        #
        # Write a single row containing the headers
        # Make the headers bold and freeze panes below that row
        #
        header_names = [name for name, type in headers]
        if optimised:
            ws.append(header_names)
        else:
            header_range = ws.range("A1:%s" % ws.cell(row=0, column=len(header_names)-1).get_coordinate())
            for cell, header_name in zip(cells_in_range(header_range), header_names):
                cell.value = header_name
                cell.style.font.bold = True
                cell.style.fill.fill_type = openpyxl.style.Fill.FILL_SOLID
                cell.style.fill.start_color.index = "ffffff80"
            ws.freeze_panes = "A2"
        yield "%s headers..." % sheet_name

        #
        # Append each row to the bottom of the sheet
        #
        n_row = 0
        for n_row, row in enumerate(rowset):
            if (1 + n_row) % ROWSET_SIZE == 0:
                yield "%s row %d" % (sheet_name, n_row)
            ws.append(list(row))
        yield "%s %d rows" % (sheet_name,  n_row)

    if callback:
        wb = callback(wb)

    wb.save(filename=spreadsheet_filepath)
